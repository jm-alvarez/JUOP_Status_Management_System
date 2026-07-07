from importlib.resources import path
import tkinter
from tkinter import ttk
from tkinter import messagebox
from tkinter import filedialog
from ttkthemes import ThemedTk
from datetime import datetime
import tkcalendar
import mysql.connector
import openpyxl
import pathlib
from pathlib import Path

file_path_str = str("No file selected")

#Function for browsing Excel file and loading data to treeview
def browse_file():
    global file_path
    file_path = filedialog.askopenfilename(filetypes = [("Excel files", "*.xlsx")])
    if file_path:
        print("Selected file:", file_path)
        messagebox.showinfo("File Selected", f"You selected: {file_path}")
        
        file_path_str = str(file_path)
        file_path_label.config(text = "Selected file: " + file_path_str)
        
        load_data()
        
    else:
        print("No file selected.")
        messagebox.showwarning("No File Selected", "Please select an Excel file to load data.")
        add_button.config(state = "disabled")
        update_button.config(state = "disabled")
        delete_button.config(state = "disabled")
    
#Load data from Excel file to treeview function
def load_data():

    path = file_path
    
    print("File Path: ", path)
    
    if not path:
        messagebox.showwarning("No File Selected", "Please select an Excel file to load data.")
        add_button.config(state = "disabled")
        update_button.config(state = "disabled")
        delete_button.config(state = "disabled")
        return
    
     
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    list_values = list(sheet.values)
    print(list_values)
    
    #Read value from Excel file and insert to treeview
    
    for col_name in list_values[0]:
        treeview.heading(col_name, text = col_name)
        
    #for row in sheet.iter_rows():
    #    for cell in row:
    #        if isinstance(cell.value, datetime):
    #            cell.value = cell.value.strftime("%m/%d/%Y")
   
    for value_tuple in list_values[1:]:
        treeview.insert('', tkinter.END, values = value_tuple)
    
    add_button.config(state = "normal")
    delete_button.config(state = "normal")
        
#Insert row to Excel file and treeview function
def insert_row():
    date = formatted_date
    juop_partner = juop_partner_entry.get()
    inspection_report_date = formatted_inspection_report_date
    coverage_area = coverage_area_entry.get()
    pole_downguy_duple = number_of_poles_entry.get() + "/" + number_of_downguys_entry.get()
    remarks_duple = str("GM-") + remarks_entry.get()
    
    print("Date:", date, "\nJUOP Partner:", juop_partner, "\nInspection Report Date:", inspection_report_date, "\nCoverage Area:", coverage_area, "\nNO. OF POLES/DOWNGUYS", pole_downguy_duple, "\nRemarks:", remarks_duple)

    #Insert row to Excel file
    path = Path("C:/Users/John Mark/Documents/GitHub/JUOP_Status_Management_System/juop_data.xlsx")
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    row_values = [date, juop_partner, inspection_report_date, coverage_area, pole_downguy_duple, remarks_duple]
    sheet.append(row_values)
    workbook.save(path)
    treeview.insert('', tkinter.END, values = row_values)
    messagebox.showinfo("Success", "Data added successfully!")
    
    #Clear input fields after adding data
    date_entry.set_date(datetime.now())
    juop_partner_entry.delete(0, tkinter.END)
    inspection_report_date_entry.set_date(datetime.now())
    coverage_area_entry.delete(0, tkinter.END)
    number_of_poles_entry.delete(0, tkinter.END)
    number_of_downguys_entry.delete(0, tkinter.END)
    remarks_entry.delete(0, tkinter.END)

def update_row():
    path = file_path
    
def colnum_to_letter(n):
    letters = ""
    while n > 0:
        n, remainder = divmod(n-1, 26)
        letters = chr(65 + remainder) + letters
    return letters

window = ThemedTk(theme = "breeze")
window.title("JUOP Status Management System")
window.state("zoomed")
window.resizable(False, False)

screen_height = window.winfo_screenheight()
screen_width = window.winfo_screenwidth()

frame = ttk.Frame(window)
frame.pack()

add_data_frame = ttk.LabelFrame(frame, text = "Add Data")
add_data_frame.grid(row = 0, column = 0)

date_label = ttk.Label(add_data_frame, text = "Date")
date_label.grid(row = 0, column = 0)
date_entry = tkcalendar.DateEntry(add_data_frame, date_pattern = "mm/dd/yyyy")
formatted_date = date_entry.get_date().strftime("%m/%d/%Y")
date_entry.grid(row = 1, column = 0)

juop_partner_label = ttk.Label(add_data_frame, text = "JUOP Partner")
juop_partner_label.grid(row = 0, column = 1)
juop_partner_entry = ttk.Entry(add_data_frame, width = 20)
juop_partner_entry.grid(row = 1, column = 1)

inspection_report_date_label = ttk.Label(add_data_frame, text = "Inspection Report Date")
inspection_report_date_label.grid(row = 3, column = 0)
inspection_report_date_entry = tkcalendar.DateEntry(add_data_frame, date_pattern = "mm/dd/yyyy")
formatted_inspection_report_date = inspection_report_date_entry.get_date().strftime("%m/%d/%Y")
inspection_report_date_entry.grid(row = 4, column = 0)

coverage_area_label = ttk.Label(add_data_frame, text = "Coverage Area")
coverage_area_label.grid(row = 3, column = 1)
coverage_area_entry = ttk.Entry(add_data_frame, width = 20)
coverage_area_entry.grid(row = 4, column = 1)

number_of_poles_label = ttk.Label(add_data_frame, text = "No. of Poles")
number_of_poles_label.grid(row = 5, column = 0)
number_of_poles_entry = ttk.Spinbox(add_data_frame, from_ = 0, to = 1000, width = 10, validate = "key", validatecommand = (add_data_frame.register(lambda s: s.isdigit()), "%P"))
number_of_poles_entry.grid(row = 6, column = 0)

number_of_downguys_label = ttk.Label(add_data_frame, text = "No. of Downguys")
number_of_downguys_label.grid(row = 5, column = 1)
number_of_downguys_entry = ttk.Spinbox(add_data_frame, from_ = 0, to = 1000, width = 10, validate = "key", validatecommand = (add_data_frame.register(lambda s: s.isdigit()), "%P"))
number_of_downguys_entry.grid(row = 6, column = 1)

remarks_frame = ttk.LabelFrame(add_data_frame, text = "Remarks")
remarks_frame.grid(row = 7, column = 0, columnspan = 2, sticky = "news", padx = 10, pady = 10)
gm_label = ttk.Label(remarks_frame, text = "GM-")
gm_label.grid(row = 0, column = 0)
#remarks_entry = ttk.Entry(remarks_frame, width = 5, validate = "key", validatecommand = (remarks_frame.register(lambda s: s.isdigit() and len(s) <= 4), "%P"))
remarks_entry = ttk.Entry(remarks_frame, width = 10)
remarks_entry.grid(row = 0, column = 1)

buttons_frame = ttk.LabelFrame(add_data_frame, text = "Actions")
buttons_frame.grid(row = 8, column = 0, columnspan = 2, sticky = "news")
add_button = ttk.Button(buttons_frame, text = "Add Data", command = insert_row)
add_button.grid(row = 0, column = 0)
add_button.config(state = "disabled")

update_button = ttk.Button(buttons_frame, text = "Update Data")
update_button.grid(row = 0, column = 1)
update_button.config(state = "disabled")

delete_button = ttk.Button(buttons_frame, text = "Delete Data")
delete_button.grid(row = 0, column = 2)
delete_button.config(state = "disabled")


file_frame = ttk.LabelFrame(add_data_frame, text = "File Path")
file_frame.grid(row = 9, column = 0, columnspan = 2, sticky = "news", padx = 20, pady = 20)
file_path_label = ttk.Label(file_frame, text = "No file selected")
file_path_label.pack()
file_path_label.config(text = "Selected file: " + file_path_str, wraplength = 300)
browse_button = ttk.Button(buttons_frame, text = "Browse File", command = browse_file)
browse_button.grid(row = 0, column = 3)


for widget in add_data_frame.winfo_children():
    widget.grid_configure(padx = 5, pady = 5)

for widget in buttons_frame.winfo_children():
    widget.grid_configure(padx = 5, pady = 5)
    
tree_frame = tkinter.LabelFrame(frame, text = "Data")
tree_frame.grid(row = 0, column = 1, pady = 20)
tree_scrollbar = tkinter.Scrollbar(tree_frame)
tree_scrollbar.pack(side = "right", fill = "y")

cols = ("DATE", "JUOP PARTNER", "INSPECTION REPORT DATE", "COVERAGE AREA", "NO. OF POLES/DOWNGUYS", "REMARKS")
treeview = tkinter.ttk.Treeview(tree_frame, show = "headings", yscrollcommand = tree_scrollbar.set,columns = cols, height = 10)
treeview.column("DATE", width = 100)
treeview.column("JUOP PARTNER", width = 150)
treeview.column("INSPECTION REPORT DATE", width = 150)
treeview.column("COVERAGE AREA", width = 350)
treeview.column("NO. OF POLES/DOWNGUYS", width = 100)
treeview.column("REMARKS", width = 100)
treeview.pack()
tree_scrollbar.config(command = treeview.yview)

window.mainloop()


#Add Payment, Update Payment, Payment Date, Payment History, and Payment Status columns to Excel file and treeview 