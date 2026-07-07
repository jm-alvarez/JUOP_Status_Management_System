"""
JUOP Management System
A GUI application for managing JUOP data stored in an Excel file.
Requirements: pip install openpyxl
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import os
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)

# ─────────────────────────────────────────────
#  Constants
# ─────────────────────────────────────────────
EXCEL_FILE = "JUOP_Data1.xlsx"
SHEET_NAME = "JUOP Records"

COLUMNS = [
    "DATE",
    "JUOP PARTNER",
    "INSPECTION REPORT DATE",
    "COVERAGE AREA",
    "NO. OF POLES/DOWNGUYS",
    "PAYMENT",
    "PAYMENT DATE",
    "UPDATE PAYMENT",
    "REMARKS",
]

# Theme colors
CLR_PRIMARY    = "#1A3C6E"   # deep navy
CLR_SECONDARY  = "#2E6DA4"   # medium blue
CLR_ACCENT     = "#F4A836"   # amber
CLR_BG         = "#F0F4F8"   # light gray-blue
CLR_WHITE      = "#FFFFFF"
CLR_ROW_ODD    = "#FFFFFF"
CLR_ROW_EVEN   = "#EAF2FB"
CLR_HEADER_BG  = "#1A3C6E"
CLR_HEADER_FG  = "#FFFFFF"
CLR_BTN_ADD    = "#27AE60"
CLR_BTN_UPD    = "#2980B9"
CLR_BTN_DEL    = "#E74C3C"
CLR_BTN_CLR    = "#7F8C8D"
CLR_BTN_EXP    = "#8E44AD"


# ─────────────────────────────────────────────
#  Excel helpers
# ─────────────────────────────────────────────
def init_excel():
    """Create the Excel file with headers if it doesn't exist."""
    if os.path.exists(EXCEL_FILE):
        return
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    header_fill = PatternFill("solid", fgColor="1A3C6E")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin        = Side(style="thin", color="AAAAAA")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    col_widths = [15, 25, 22, 22, 22, 15, 15, 18, 25]

    for col_idx, (col_name, width) in enumerate(zip(COLUMNS, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = width

    ws.row_dimensions[1].height = 35
    wb.save(EXCEL_FILE)


def load_records():
    """Return all data rows as list-of-lists (excluding header)."""
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):
            rows.append(list(row))
    return rows


def save_all_records(records):
    """Overwrite data rows in the Excel file."""
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]

    # Clear existing data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    thin   = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    for r_idx, record in enumerate(records, start=2):
        fill_color = "FFFFFF" if r_idx % 2 == 0 else "EAF2FB"
        row_fill   = PatternFill("solid", fgColor=fill_color)
        for c_idx, value in enumerate(record, start=1):
            cell           = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border    = border
            cell.fill      = row_fill
            cell.alignment = center if c_idx in (1, 3, 5, 6, 7, 8) else left
            cell.font      = Font(name="Arial", size=10)
        ws.row_dimensions[r_idx].height = 20

    wb.save(EXCEL_FILE)


# ─────────────────────────────────────────────
#  Main Application
# ─────────────────────────────────────────────
class JUOPApp(tk.Tk):
    def __init__(self):
        super().__init__()
        init_excel()

        self.title("JUOP Management System")
        self.state("zoomed")          # maximized on Windows; fallback below
        self.configure(bg=CLR_BG)
        self.minsize(1100, 650)

        try:
            self.state("zoomed")
        except Exception:
            self.geometry("1280x720")

        self.selected_row_index = None   # 0-based index in self.records
        self.records = []
        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", self._on_search)

        self._build_ui()
        self.refresh_table()

    # ──────────────────────────────────────────
    #  UI Build
    # ──────────────────────────────────────────
    def _build_ui(self):
        self._build_header()
        self._build_toolbar()
        self._build_form()
        self._build_table()
        self._build_statusbar()

    def _build_header(self):
        hdr = tk.Frame(self, bg=CLR_PRIMARY, height=65)
        hdr.pack(fill="x", side="top")
        hdr.pack_propagate(False)

        tk.Label(
            hdr,
            text="⚡  JUOP MANAGEMENT SYSTEM",
            bg=CLR_PRIMARY, fg=CLR_WHITE,
            font=("Arial", 20, "bold"),
            padx=20,
        ).pack(side="left", pady=10)

        tk.Label(
            hdr,
            text=f"File: {EXCEL_FILE}",
            bg=CLR_PRIMARY, fg="#A9C8E8",
            font=("Arial", 10),
            padx=20,
        ).pack(side="right", pady=10)

    def _build_toolbar(self):
        bar = tk.Frame(self, bg=CLR_SECONDARY, height=42)
        bar.pack(fill="x", side="top")
        bar.pack_propagate(False)

        btn_cfg = [
            ("➕  Add Record",    CLR_BTN_ADD, self.add_record),
            ("✏️  Update Record", CLR_BTN_UPD, self.update_record),
            ("🗑️  Delete Record", CLR_BTN_DEL, self.delete_record),
            ("🔄  Clear Form",    CLR_BTN_CLR, self.clear_form),
            ("📤  Export Excel",  CLR_BTN_EXP, self.export_copy),
        ]
        for label, color, cmd in btn_cfg:
            tk.Button(
                bar, text=label, command=cmd,
                bg=color, fg=CLR_WHITE,
                font=("Arial", 10, "bold"),
                relief="flat", padx=12, pady=6,
                cursor="hand2", activebackground=color,
                activeforeground=CLR_WHITE,
                bd=0,
            ).pack(side="left", padx=6, pady=5)

        # Search
        tk.Label(
            bar, text="🔍 Search:", bg=CLR_SECONDARY, fg=CLR_WHITE,
            font=("Arial", 10, "bold"),
        ).pack(side="right", padx=(0, 4), pady=5)
        tk.Entry(
            bar, textvariable=self.search_var,
            font=("Arial", 10), width=22,
            relief="flat", bd=4,
        ).pack(side="right", padx=(0, 10), pady=5)

    def _build_form(self):
        outer = tk.LabelFrame(
            self, text="  Entry Form  ",
            font=("Arial", 10, "bold"),
            bg=CLR_BG, fg=CLR_PRIMARY,
            bd=2, relief="groove", padx=10, pady=8,
        )
        outer.pack(fill="x", padx=12, pady=(8, 4))

        self.entries: dict[str, tk.Widget] = {}

        # Layout: 3 rows × 3 columns of fields
        field_layout = [
            [COLUMNS[0], COLUMNS[1], COLUMNS[2]],
            [COLUMNS[3], COLUMNS[4], COLUMNS[5]],
            [COLUMNS[6], COLUMNS[7], COLUMNS[8]],
        ]

        for r, row_fields in enumerate(field_layout):
            for c, field in enumerate(row_fields):
                frame = tk.Frame(outer, bg=CLR_BG)
                frame.grid(row=r, column=c, padx=10, pady=4, sticky="ew")
                outer.columnconfigure(c, weight=1)

                tk.Label(
                    frame, text=field, bg=CLR_BG, fg=CLR_PRIMARY,
                    font=("Arial", 9, "bold"), anchor="w",
                ).pack(fill="x")

                if field == "REMARKS":
                    widget = tk.Text(
                        frame, height=2, width=28,
                        font=("Arial", 10), relief="solid", bd=1,
                        wrap="word",
                    )
                else:
                    widget = tk.Entry(
                        frame, font=("Arial", 10),
                        relief="solid", bd=1, width=28,
                    )
                widget.pack(fill="x", ipady=3)
                self.entries[field] = widget

    def _build_table(self):
        container = tk.Frame(self, bg=CLR_BG)
        container.pack(fill="both", expand=True, padx=12, pady=(4, 4))

        # Record count label
        self.count_label = tk.Label(
            container, text="Records: 0",
            bg=CLR_BG, fg=CLR_PRIMARY,
            font=("Arial", 9, "bold"), anchor="w",
        )
        self.count_label.pack(anchor="w")

        # Treeview + scrollbars
        tree_frame = tk.Frame(container, bg=CLR_BG)
        tree_frame.pack(fill="both", expand=True)

        style = ttk.Style()
        style.theme_use("clam")
        style.configure(
            "JUOP.Treeview",
            background=CLR_ROW_ODD,
            foreground="#222222",
            fieldbackground=CLR_ROW_ODD,
            rowheight=24,
            font=("Arial", 9),
        )
        style.configure(
            "JUOP.Treeview.Heading",
            background=CLR_HEADER_BG,
            foreground=CLR_HEADER_FG,
            font=("Arial", 9, "bold"),
            relief="flat",
        )
        style.map(
            "JUOP.Treeview",
            background=[("selected", CLR_SECONDARY)],
            foreground=[("selected", CLR_WHITE)],
        )

        self.tree = ttk.Treeview(
            tree_frame,
            columns=COLUMNS,
            show="headings",
            style="JUOP.Treeview",
            selectmode="browse",
        )

        col_widths = [90, 150, 130, 140, 150, 100, 100, 120, 160]
        for col, width in zip(COLUMNS, col_widths):
            self.tree.heading(col, text=col, anchor="center")
            self.tree.column(col, width=width, anchor="center", minwidth=80)

        # Alternating row tags
        self.tree.tag_configure("odd",  background=CLR_ROW_ODD)
        self.tree.tag_configure("even", background=CLR_ROW_EVEN)

        vsb = ttk.Scrollbar(tree_frame, orient="vertical",   command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal",  command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_frame.rowconfigure(0, weight=1)
        tree_frame.columnconfigure(0, weight=1)

        self.tree.bind("<<TreeviewSelect>>", self._on_row_select)

    def _build_statusbar(self):
        self.status_var = tk.StringVar(value="Ready")
        sb = tk.Label(
            self, textvariable=self.status_var,
            bg=CLR_PRIMARY, fg=CLR_WHITE,
            font=("Arial", 9), anchor="w", padx=10,
        )
        sb.pack(fill="x", side="bottom")

    # ──────────────────────────────────────────
    #  Helpers
    # ──────────────────────────────────────────
    def _get_form_values(self):
        values = []
        for field in COLUMNS:
            widget = self.entries[field]
            if isinstance(widget, tk.Text):
                val = widget.get("1.0", "end-1c").strip()
            else:
                val = widget.get().strip()
            values.append(val if val else None)
        return values

    def _populate_form(self, record):
        for field, value in zip(COLUMNS, record):
            widget = self.entries[field]
            if isinstance(widget, tk.Text):
                widget.delete("1.0", "end")
                widget.insert("1.0", value or "")
            else:
                widget.delete(0, "end")
                widget.insert(0, value or "")

    def _set_status(self, msg, color=CLR_PRIMARY):
        self.status_var.set(f"  {msg}")
        self.configure(bg=CLR_BG)  # keep bg stable

    def refresh_table(self, filter_text=""):
        self.records = load_records()
        self.tree.delete(*self.tree.get_children())
        shown = 0
        ft = filter_text.lower()
        for i, rec in enumerate(self.records):
            if ft and not any(ft in str(v).lower() for v in rec if v):
                continue
            tag = "even" if shown % 2 == 0 else "odd"
            self.tree.insert("", "end", iid=str(i), values=rec, tags=(tag,))
            shown += 1
        self.count_label.config(text=f"Records: {shown}  (Total: {len(self.records)})")
        self.selected_row_index = None
        self._set_status(f"Loaded {len(self.records)} records.")

    def _on_row_select(self, event):
        sel = self.tree.selection()
        if not sel:
            return
        iid = sel[0]
        self.selected_row_index = int(iid)
        record = self.records[self.selected_row_index]
        self._populate_form(record)
        self._set_status(f"Selected row {self.selected_row_index + 1}.")

    def _on_search(self, *_):
        self.refresh_table(self.search_var.get())

    # ──────────────────────────────────────────
    #  CRUD Operations
    # ──────────────────────────────────────────
    def add_record(self):
        values = self._get_form_values()
        if not any(values):
            messagebox.showwarning("Empty Form", "Please fill in at least one field.")
            return
        self.records.append(values)
        save_all_records(self.records)
        self.clear_form()
        self.refresh_table(self.search_var.get())
        self._set_status("✅ Record added successfully.")

    def update_record(self):
        if self.selected_row_index is None:
            messagebox.showwarning("No Selection", "Please select a record from the table to update.")
            return
        values = self._get_form_values()
        if not any(values):
            messagebox.showwarning("Empty Form", "Please fill in at least one field.")
            return
        self.records[self.selected_row_index] = values
        save_all_records(self.records)
        self.clear_form()
        self.refresh_table(self.search_var.get())
        self._set_status("✏️ Record updated successfully.")

    def delete_record(self):
        if self.selected_row_index is None:
            messagebox.showwarning("No Selection", "Please select a record from the table to delete.")
            return
        rec = self.records[self.selected_row_index]
        preview = str(rec[1] or rec[0] or "this record")
        if not messagebox.askyesno(
            "Confirm Delete",
            f"Are you sure you want to delete:\n\n{preview}?\n\nThis action cannot be undone.",
        ):
            return
        del self.records[self.selected_row_index]
        save_all_records(self.records)
        self.clear_form()
        self.refresh_table(self.search_var.get())
        self._set_status("🗑️ Record deleted.")

    def clear_form(self):
        for field in COLUMNS:
            widget = self.entries[field]
            if isinstance(widget, tk.Text):
                widget.delete("1.0", "end")
            else:
                widget.delete(0, "end")
        self.selected_row_index = None
        self.tree.selection_remove(self.tree.selection())
        self._set_status("Form cleared.")

    def export_copy(self):
        dest = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"JUOP_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            title="Export a copy of JUOP Data",
        )
        if not dest:
            return
        import shutil
        shutil.copy2(EXCEL_FILE, dest)
        messagebox.showinfo("Export Successful", f"File exported to:\n{dest}")
        self._set_status(f"📤 Exported to {os.path.basename(dest)}")


# ─────────────────────────────────────────────
#  Entry Point
# ─────────────────────────────────────────────
if __name__ == "__main__":
    app = JUOPApp()
    app.mainloop()