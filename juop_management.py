import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
import os
from datetime import datetime, date
import copy

# ── Constants ────────────────────────────────────────────────────────────────

COLUMNS = [
    "DATE",
    "JUOP PARTNER",
    "INSPECTION REPORT DATE",
    "COVERAGE AREA",
    "NO. OF POLES/DOWNGUYS",
    "PAYMENT",
    "PAYMENT DATE",
    "UPDATE PAYMENT",
    "REMARKS",
]

COL_WIDTHS = [14, 22, 22, 20, 22, 14, 14, 16, 22]

EXCEL_FILE = "JUOP_Management.xlsx"
SHEET_NAME = "JUOP Records"

# ── Colour palette ────────────────────────────────────────────────────────────

BG_DARK      = "#0D1B2A"   # navy
BG_MID       = "#1B2A3B"
BG_PANEL     = "#162032"
ACCENT_BLUE  = "#1E88E5"
ACCENT_LIGHT = "#42A5F5"
ACCENT_GREEN = "#00C853"
ACCENT_RED   = "#FF5252"
ACCENT_AMBER = "#FFB300"
TEXT_WHITE   = "#FFFFFF"
TEXT_LIGHT   = "#B0BEC5"
TEXT_DARK    = "#0D1B2A"
ROW_ODD      = "#1B2A3B"
ROW_EVEN     = "#162032"
ROW_SEL      = "#1E3A5F"
HEADER_BG    = "#1E88E5"

# ── Excel helpers ─────────────────────────────────────────────────────────────

def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def style_header(ws):
    hdr_fill = PatternFill("solid", fgColor="1E88E5")
    hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = hdr_align
        cell.border = thin_border()
    ws.row_dimensions[1].height = 32

def style_data_row(ws, row_num):
    fill_color = "EAF4FB" if row_num % 2 == 0 else "FFFFFF"
    fill = PatternFill("solid", fgColor=fill_color)
    font = Font(name="Arial", size=10, color="0D1B2A")
    align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = thin_border()

def init_excel():
    if os.path.exists(EXCEL_FILE):
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    style_header(ws)
    for i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    wb.save(EXCEL_FILE)

def load_records():
    init_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell is not None and str(cell).strip() != "" for cell in row):
            records.append(list(row))
    return records

def save_records(records):
    init_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    # Clear old data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None
    # Write fresh data
    for r_idx, record in enumerate(records, start=2):
        for c_idx, value in enumerate(record, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
        style_data_row(ws, r_idx)
    wb.save(EXCEL_FILE)

# ── Tkinter helpers ───────────────────────────────────────────────────────────

def make_btn(parent, text, command, color=ACCENT_BLUE, fg=TEXT_WHITE,
             width=14, font_size=9):
    btn = tk.Button(
        parent, text=text, command=command,
        bg=color, fg=fg, activebackground=color,
        activeforeground=fg, relief="flat", cursor="hand2",
        font=("Segoe UI", font_size, "bold"), width=width,
        pady=6, padx=4
    )
    btn.bind("<Enter>", lambda e: btn.config(bg=_lighten(color)))
    btn.bind("<Leave>", lambda e: btn.config(bg=color))
    return btn

def _lighten(hex_color):
    hex_color = hex_color.lstrip("#")
    r, g, b = (int(hex_color[i:i+2], 16) for i in (0, 2, 4))
    r = min(255, r + 30)
    g = min(255, g + 30)
    b = min(255, b + 30)
    return f"#{r:02X}{g:02X}{b:02X}"

# ── Entry / Edit Dialog ───────────────────────────────────────────────────────

class RecordDialog(tk.Toplevel):
    """Modal dialog for Add / Edit."""

    def __init__(self, parent, title="Add Record", initial=None):
        super().__init__(parent)
        self.title(title)
        self.configure(bg=BG_DARK)
        self.resizable(False, False)
        self.grab_set()
        self.result = None

        self.entries = {}
        self._build(initial or [""] * len(COLUMNS))
        self.update_idletasks()
        self._center(parent)

    def _center(self, parent):
        pw = parent.winfo_rootx()
        py = parent.winfo_rooty()
        ph = parent.winfo_height()
        pw_w = parent.winfo_width()
        w = self.winfo_width()
        h = self.winfo_height()
        x = pw + (pw_w - w) // 2
        y = py + (ph - h) // 2
        self.geometry(f"+{x}+{y}")

    def _build(self, initial):
        # Title bar
        hdr = tk.Frame(self, bg=ACCENT_BLUE, pady=10)
        hdr.pack(fill="x")
        tk.Label(hdr, text=self.title(), bg=ACCENT_BLUE, fg=TEXT_WHITE,
                 font=("Segoe UI", 13, "bold")).pack()

        body = tk.Frame(self, bg=BG_DARK, padx=20, pady=10)
        body.pack(fill="both")

        date_cols = {"DATE", "INSPECTION REPORT DATE", "PAYMENT DATE"}

        for i, col in enumerate(COLUMNS):
            tk.Label(body, text=col, bg=BG_DARK, fg=TEXT_LIGHT,
                     font=("Segoe UI", 9, "bold"),
                     anchor="w").grid(row=i, column=0, sticky="w",
                                      pady=4, padx=(0, 12))
            val = initial[i] if initial[i] is not None else ""
            if isinstance(val, (datetime, date)):
                val = val.strftime("%Y-%m-%d")

            ent = tk.Entry(body, bg=BG_MID, fg=TEXT_WHITE,
                           insertbackground=TEXT_WHITE,
                           relief="flat", font=("Segoe UI", 10),
                           width=34,
                           highlightthickness=1,
                           highlightcolor=ACCENT_BLUE,
                           highlightbackground="#2A3A4A")
            ent.insert(0, str(val))
            ent.grid(row=i, column=1, pady=4)
            if col in date_cols:
                tk.Label(body, text="(YYYY-MM-DD)", bg=BG_DARK,
                         fg="#607D8B", font=("Segoe UI", 8)).grid(
                    row=i, column=2, padx=(6, 0))
            self.entries[col] = ent

        # Buttons
        btn_frame = tk.Frame(self, bg=BG_DARK, pady=14)
        btn_frame.pack()
        make_btn(btn_frame, "✔  Save", self._save,
                 color=ACCENT_GREEN, fg=TEXT_DARK).pack(side="left", padx=8)
        make_btn(btn_frame, "✖  Cancel", self.destroy,
                 color=ACCENT_RED, fg=TEXT_WHITE).pack(side="left", padx=8)

    def _save(self):
        values = []
        for col in COLUMNS:
            values.append(self.entries[col].get().strip())
        self.result = values
        self.destroy()

# ── Main Application ──────────────────────────────────────────────────────────

class JUOPApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("JUOP Management System")
        self.state("zoomed")
        self.configure(bg=BG_DARK)
        self.minsize(1100, 600)

        self.records = []
        self._sort_col = None
        self._sort_asc = True
        self._search_var = tk.StringVar()
        self._search_var.trace_add("write", lambda *_: self._refresh_tree())

        self._build_ui()
        self._load()

    # ── UI ────────────────────────────────────────────────────────────────────

    def _build_ui(self):
        self._build_topbar()
        self._build_toolbar()
        self._build_search()
        self._build_tree()
        self._build_statusbar()

    def _build_topbar(self):
        bar = tk.Frame(self, bg=ACCENT_BLUE, height=54)
        bar.pack(fill="x")
        bar.pack_propagate(False)
        tk.Label(bar, text="⚡  JUOP MANAGEMENT SYSTEM",
                 bg=ACCENT_BLUE, fg=TEXT_WHITE,
                 font=("Segoe UI", 16, "bold")).pack(side="left", padx=18)
        tk.Label(bar, text=f"File: {EXCEL_FILE}", bg=ACCENT_BLUE,
                 fg="#BBDEFB", font=("Segoe UI", 9)).pack(side="right", padx=16)

    def _build_toolbar(self):
        bar = tk.Frame(self, bg=BG_MID, pady=8)
        bar.pack(fill="x", padx=0)

        btns = [
            ("➕  Add Record",    self._add,    ACCENT_GREEN,  TEXT_DARK),
            ("✏️  Edit Record",   self._edit,   ACCENT_BLUE,   TEXT_WHITE),
            ("🗑️  Delete Record", self._delete, ACCENT_RED,    TEXT_WHITE),
            ("🔄  Refresh",       self._load,   ACCENT_AMBER,  TEXT_DARK),
            ("💾  Export Copy",   self._export, "#7B1FA2",     TEXT_WHITE),
        ]
        for text, cmd, color, fg in btns:
            make_btn(bar, text, cmd, color=color, fg=fg,
                     width=16).pack(side="left", padx=8)

        # Sort controls (right side)
        sort_frame = tk.Frame(bar, bg=BG_MID)
        sort_frame.pack(side="right", padx=12)
        tk.Label(sort_frame, text="Sort by:", bg=BG_MID, fg=TEXT_LIGHT,
                 font=("Segoe UI", 9)).pack(side="left")
        self._sort_var = tk.StringVar(value=COLUMNS[0])
        cb = ttk.Combobox(sort_frame, textvariable=self._sort_var,
                          values=COLUMNS, state="readonly", width=22,
                          font=("Segoe UI", 9))
        cb.pack(side="left", padx=6)

        self._asc_var = tk.BooleanVar(value=True)
        tk.Checkbutton(sort_frame, text="Ascending",
                       variable=self._asc_var,
                       bg=BG_MID, fg=TEXT_LIGHT,
                       selectcolor=BG_MID,
                       activebackground=BG_MID,
                       font=("Segoe UI", 9)).pack(side="left")
        make_btn(sort_frame, "Sort ▼", self._sort,
                 color=ACCENT_BLUE, width=8).pack(side="left", padx=6)

    def _build_search(self):
        sbar = tk.Frame(self, bg=BG_PANEL, pady=6)
        sbar.pack(fill="x")
        tk.Label(sbar, text="🔍 Search:", bg=BG_PANEL, fg=TEXT_LIGHT,
                 font=("Segoe UI", 9)).pack(side="left", padx=(14, 4))
        se = tk.Entry(sbar, textvariable=self._search_var,
                      bg=BG_MID, fg=TEXT_WHITE,
                      insertbackground=TEXT_WHITE,
                      relief="flat", font=("Segoe UI", 10), width=42,
                      highlightthickness=1,
                      highlightcolor=ACCENT_BLUE,
                      highlightbackground="#2A3A4A")
        se.pack(side="left", padx=4)
        tk.Label(sbar, text="(searches all columns)",
                 bg=BG_PANEL, fg="#607D8B",
                 font=("Segoe UI", 8)).pack(side="left", padx=4)

    def _build_tree(self):
        container = tk.Frame(self, bg=BG_DARK)
        container.pack(fill="both", expand=True, padx=10, pady=(4, 0))

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("JUOP.Treeview",
                         background=ROW_ODD,
                         foreground=TEXT_WHITE,
                         fieldbackground=ROW_ODD,
                         rowheight=26,
                         font=("Segoe UI", 9))
        style.configure("JUOP.Treeview.Heading",
                         background=HEADER_BG,
                         foreground=TEXT_WHITE,
                         font=("Segoe UI", 9, "bold"),
                         relief="flat")
        style.map("JUOP.Treeview",
                  background=[("selected", ROW_SEL)],
                  foreground=[("selected", TEXT_WHITE)])
        style.map("JUOP.Treeview.Heading",
                  background=[("active", ACCENT_LIGHT)])

        self.tree = ttk.Treeview(container, columns=COLUMNS,
                                 show="headings", style="JUOP.Treeview",
                                 selectmode="browse")

        col_widths_px = [100, 140, 140, 130, 140, 100, 100, 120, 150]
        for col, w in zip(COLUMNS, col_widths_px):
            self.tree.heading(col, text=col,
                              command=lambda c=col: self._col_sort(c))
            self.tree.column(col, width=w, minwidth=60, anchor="w")

        # Scrollbars
        vsb = ttk.Scrollbar(container, orient="vertical",
                            command=self.tree.yview)
        hsb = ttk.Scrollbar(container, orient="horizontal",
                            command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set,
                            xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        self.tree.tag_configure("odd",  background=ROW_ODD)
        self.tree.tag_configure("even", background=ROW_EVEN)
        self.tree.bind("<Double-1>", lambda e: self._edit())

    def _build_statusbar(self):
        self.status_var = tk.StringVar(value="Ready")
        bar = tk.Frame(self, bg="#0A1520", height=26)
        bar.pack(fill="x", side="bottom")
        bar.pack_propagate(False)
        tk.Label(bar, textvariable=self.status_var,
                 bg="#0A1520", fg=TEXT_LIGHT,
                 font=("Segoe UI", 9), anchor="w").pack(
            side="left", padx=10)

    # ── Data helpers ──────────────────────────────────────────────────────────

    def _load(self):
        self.records = load_records()
        self._refresh_tree()
        self._set_status(f"Loaded {len(self.records)} record(s) from {EXCEL_FILE}")

    def _refresh_tree(self):
        self.tree.delete(*self.tree.get_children())
        q = self._search_var.get().lower().strip()
        shown = 0
        for i, rec in enumerate(self.records):
            row = [str(v) if v is not None else "" for v in rec]
            if q and not any(q in cell.lower() for cell in row):
                continue
            tag = "even" if shown % 2 == 0 else "odd"
            self.tree.insert("", "end", iid=str(i),
                             values=row, tags=(tag,))
            shown += 1
        self._set_status(
            f"Showing {shown} of {len(self.records)} record(s) | "
            f"File: {EXCEL_FILE}")

    def _selected_index(self):
        sel = self.tree.selection()
        if not sel:
            return None
        return int(sel[0])

    def _set_status(self, msg):
        ts = datetime.now().strftime("%H:%M:%S")
        self.status_var.set(f"[{ts}]  {msg}")

    # ── CRUD ──────────────────────────────────────────────────────────────────

    def _add(self):
        dlg = RecordDialog(self, title="Add New Record")
        self.wait_window(dlg)
        if dlg.result is None:
            return
        self.records.append(dlg.result)
        save_records(self.records)
        self._refresh_tree()
        self._set_status("✔  New record added and saved.")

    def _edit(self):
        idx = self._selected_index()
        if idx is None:
            messagebox.showwarning("No Selection",
                                   "Please select a record to edit.",
                                   parent=self)
            return
        dlg = RecordDialog(self, title="Edit Record",
                           initial=self.records[idx])
        self.wait_window(dlg)
        if dlg.result is None:
            return
        self.records[idx] = dlg.result
        save_records(self.records)
        self._refresh_tree()
        self._set_status(f"✔  Record #{idx + 1} updated and saved.")

    def _delete(self):
        idx = self._selected_index()
        if idx is None:
            messagebox.showwarning("No Selection",
                                   "Please select a record to delete.",
                                   parent=self)
            return
        rec = self.records[idx]
        preview = str(rec[1]) if rec[1] else f"Row {idx + 1}"
        if not messagebox.askyesno(
                "Confirm Delete",
                f"Delete record for '{preview}'?\nThis cannot be undone.",
                parent=self):
            return
        self.records.pop(idx)
        save_records(self.records)
        self._refresh_tree()
        self._set_status(f"🗑️  Record deleted. {len(self.records)} records remaining.")

    # ── Sort ──────────────────────────────────────────────────────────────────

    def _sort(self):
        col_name = self._sort_var.get()
        asc = self._asc_var.get()
        self._do_sort(col_name, asc)

    def _col_sort(self, col_name):
        if self._sort_col == col_name:
            self._sort_asc = not self._sort_asc
        else:
            self._sort_col = col_name
            self._sort_asc = True
        self._do_sort(col_name, self._sort_asc)

    def _do_sort(self, col_name, asc):
        col_idx = COLUMNS.index(col_name)

        def sort_key(rec):
            val = rec[col_idx]
            if val is None:
                return ("", )
            s = str(val).strip()
            # Try numeric
            try:
                return (0, float(s))
            except ValueError:
                pass
            # Try date
            for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"):
                try:
                    return (0, datetime.strptime(s, fmt).timestamp())
                except ValueError:
                    pass
            return (1, s.lower())

        self.records.sort(key=sort_key, reverse=not asc)
        save_records(self.records)
        self._refresh_tree()
        direction = "ascending" if asc else "descending"
        self._set_status(f"↕  Sorted by '{col_name}' ({direction}) and saved.")

    # ── Export ────────────────────────────────────────────────────────────────

    def _export(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"JUOP_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            parent=self)
        if not path:
            return

        import shutil
        shutil.copy2(EXCEL_FILE, path)
        self._set_status(f"💾  Exported to: {path}")
        messagebox.showinfo("Export Complete",
                            f"File saved to:\n{path}", parent=self)


# ── Entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    init_excel()
    app = JUOPApp()
    app.mainloop()
